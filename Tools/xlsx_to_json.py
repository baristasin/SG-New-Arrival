#!/usr/bin/env python3
"""
Converts the student profiles spreadsheet into Assets/Game/Data/Source/students.json,
the file consumed by the Unity editor tool "Tools/Student Data/Generate Databases from JSON"
(StudentDataImporter.cs) which fills StudentDatabase.asset.

Pipeline:  xlsx  ->  [this script]  ->  students.json  ->  [Unity importer]  ->  StudentDatabase.asset

Enums are emitted as integers because Unity's JsonUtility deserializes enums by value.
Dates are emitted as "yyyy-MM-dd" strings (empty when blank), matching StudentProfile's convention.

Usage:
    python3 Tools/xlsx_to_json.py
    python3 Tools/xlsx_to_json.py path/to/other.xlsx
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = REPO / "Assets" / "student_profiles_germany_updated.xlsx"
OUT_JSON = REPO / "Assets" / "Game" / "Data" / "Source" / "students.json"
SHEET = "Student Profiles"

# Enum value maps (must match the C# enums in Game/Scripts/StudentData/Enums.cs).
LANG = {"none": 0, "a1": 1, "a2": 2, "b1": 3, "b2": 4, "c1": 5, "c2": 6, "fluent": 7, "native": 8}
GENDER = {"male": 0, "female": 1, "diverse": 2}
MARITAL = {"single": 0, "married": 1, "divorced": 2, "widowed": 3}
BUDGET = {"low": 0, "medium": 1, "high": 2}
PHOTO_QUALITY = {"normal": 1, "sunglass": 2, "sunglasses": 2, "blurred": 3}  # blank/None -> 0
CHECK = {"no": 0, "yes": 1, "not required": 2}

warnings = []


def s(v):
    return "" if v is None else str(v).strip()


def date_iso(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def to_int(v):
    return int(v) if v is not None else 0


def to_float(v):
    return float(v) if v is not None else 0.0


def kebab(first, last):
    out = re.sub(r"[^a-z0-9-]", "", f"{first}-{last}".lower().replace(" ", "-"))
    return re.sub(r"-+", "-", out).strip("-")


def enum(value, table, field, default=0, allow_blank=True):
    if value is None or s(value) == "":
        if not allow_blank:
            warnings.append(f"{field}: blank value, defaulting to {default}")
        return default
    key = s(value).lower()
    if key not in table:
        warnings.append(f"{field}: unknown value '{value}', defaulting to {default}")
        return default
    return table[key]


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[SHEET]

    items = []
    for r in range(2, ws.max_row + 1):
        def c(col):  # 1-based column index -> cell value
            return ws.cell(row=r, column=col).value

        if c(3) is None and c(4) is None:  # no first/last name -> skip empty row
            continue

        first, last = s(c(3)), s(c(4))
        student = {
            "IdNumber": to_int(c(1)),
            "Id": kebab(first, last),
            "FirstName": first,
            "LastName": last,
            "Nationality": s(c(5)),
            "PlaceOfBirth": s(c(6)),
            "DateOfBirth": date_iso(c(7)),
            "GermanLevel": enum(c(8), LANG, "GermanLevel", allow_blank=False),
            "EnglishLevel": enum(c(9), LANG, "EnglishLevel", allow_blank=False),
            "Gender": enum(c(10), GENDER, "Gender", allow_blank=False),
            "MaritalStatus": enum(c(11), MARITAL, "MaritalStatus", allow_blank=False),
            "Budget": enum(c(12), BUDGET, "Budget", allow_blank=False),
            "VisaStatus": 0,  # not in spreadsheet; default LongTerm
            "IsEnrolled": s(c(13)).lower() == "enrolled",
            "AddressInGermany": s(c(14)),
            "MoveInDate": date_iso(c(15)),
            "Wohnungsgeber": s(c(16)),
            "FormerAddressAbroad": s(c(17)),
            "HasPreviousSchufa": s(c(18)).lower() == "yes",
            "Visa": {
                "PassportId": s(c(21)),
                "PassportIssued": date_iso(c(20)),
                "PassportExpiryDate": date_iso(c(19)),
                "AdmissionDeadline": date_iso(c(22)),
                "FinancialFunds": to_int(c(23)),
                "MonthlyRelease": to_int(c(24)),
                "Iban": s(c(25)),
                "LanguageCertLevel": enum(c(26), LANG, "LanguageCertLevel"),
                "LanguageCertDate": date_iso(c(27)),
                "LanguageCertProvider": s(c(28)),
                "ApsName": s(c(29)),
                "VisaPaymentOrder": to_float(c(30)),
                "TravelInsurancePerson": s(c(31)),
                "PhotosTotal": to_int(c(32)),
                "PhotosQuality": enum(c(33), PHOTO_QUALITY, "PhotosQuality"),
                "CheckPassport": enum(c(34), CHECK, "CheckPassport"),
                "CheckLetterOfAdmission": enum(c(35), CHECK, "CheckLetterOfAdmission"),
                "CheckBlockedAccount": enum(c(36), CHECK, "CheckBlockedAccount"),
                "CheckLanguageCertificate": enum(c(37), CHECK, "CheckLanguageCertificate"),
                "CheckAps": enum(c(38), CHECK, "CheckAps"),
                "CheckVisaPaymentOrder": enum(c(39), CHECK, "CheckVisaPaymentOrder"),
                "CheckTravelInsurance": enum(c(40), CHECK, "CheckTravelInsurance"),
                "CheckBiometricPhotos": enum(c(41), CHECK, "CheckBiometricPhotos"),
            },
        }
        items.append(student)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump({"Items": items}, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Wrote {len(items)} students -> {OUT_JSON.relative_to(REPO)}")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in dict.fromkeys(warnings):
            print(f"  - {w}")


if __name__ == "__main__":
    main()
